#!/bin/python3
# coding=utf-8
#Author: yiouejv
#Email: yiouejv@126.com
#Time: 2022-05-17 14:23:05
#Name: lang_merge.py

import xlrd
import os
import re
import time
import openpyxl
from openpyxl.styles import PatternFill


raw_path = u"xls/lang/D多语言配置表.xlsx"
lang_path = u"D多语言配置表-翻译回来.xlsx"
export_path = u"多语言待翻译导出.xlsx"
lang_dir = "xls/lang"

global write_count
write_count = 0

global languageList
languageList = []

global wait_write_list
wait_write_list = []

global export_row_index
export_row_index = 4

# lang 表: 翻译回来的表
# raw 表: 本地的多语言表

def get_cell_index_by_name(sheet, row_index, colname):
    ncol = sheet.ncols
    for i in range(0, ncol):
        v = sheet.cell_value(row_index, i)
        if v == colname:
            return i
    return -1


def get_row_index_by_name(sheet, col_index, rowname):
    nrow = sheet.nrows
    for i in range(4, nrow):
        v = sheet.cell_value(i, col_index)
        if v == rowname:
            return i
    return -1


def get_key_value(sheet, key, colname):
    row = get_row_index_by_name(sheet, 0, key)
    if row == -1:
        # raise Exception("not find {} row".format(key))
        return "", -1, -1
    col = get_cell_index_by_name(sheet, 1, colname)
    if col == -1:
        # raise Exception(u"not find zhcn column")
        return "", row, -1
    return sheet.cell_value(row, col), row, col


def is_key_diff(raw_sheet, lang_sheet, key):
    colname = "zhcn"
    rv, _, _ = get_key_value(raw_sheet, key, colname)
    lv, _, _ = get_key_value(lang_sheet, key, colname)
    return rv != lv


def write_equal_list(xlspath, sheetname):
    global write_count
    global wait_write_list
    print("write_count: {}".format(write_count))
    book = openpyxl.load_workbook(xlspath)
    sheet = None
    for shet in book.worksheets:
        if shet.title == sheetname:
            sheet = shet
            break
    if not sheet:
        raise Exception("not find {} sheet".format(sheetname))

    for tup in wait_write_list:
        sheet.cell(row=tup[0], column=tup[1], value=tup[2])

    wait_write_list = []
    book.save(xlspath)


def write_key_value(xlspath, sheetname, row, col, value):
    global write_count
    global wait_write_list
    write_count += 1

    wait_write_list.append((row+1, col+1, value))
    if len(wait_write_list) >= 1000:
        write_equal_list(xlspath, sheetname)


def sync_equal_key(raw_sheet, lang_sheet, key, xlspath):
    # print("sync_equal_key, key:", key)
    for lang in languageList:
        _, row, col = get_key_value(raw_sheet, key, lang)
        lval, _, _ = get_key_value(lang_sheet, key, lang)
        # print(key, row, col)
        assert(row != -1)  # "not find key"
        assert(col != -1)  # "not find language column: ".format(lang)
        if lval != "":
            # print(u"lang: {}, raw value: {}, lang value: {}".format(lang, rval, lval))
            write_key_value(xlspath, raw_sheet.name, row, col, lval)

def get_openpyxl_sheet_by_name(book, sheetname):
    sheet = None
    for shet in book.worksheets:
        if shet.title == sheetname:
            sheet = shet
            break
    if not sheet:
        raise Exception("not find skill sheet")
    return sheet


def raw_key_fill_color(sheet, key, xlspath):
    row = get_row_index_by_name(sheet, 0, key)

    book = openpyxl.load_workbook(xlspath)
    s = get_openpyxl_sheet_by_name(book, "language")
    cell = s.cell(row=row+1, column=1)

    fill = PatternFill("solid", fgColor="ffc7ce")
    cell.fill = fill


def get_index_from_language_list(lang):
    for i, language in enumerate(languageList):
        if language == lang:
            return i + 1  # 1: pre is uid


def export_diff_key(export_sheet, raw_sheet, key):
    print("export_diff_key, key:", key)
    global export_row_index
    is_first = True
    for lang in languageList:
        rval, _, _ = get_key_value(raw_sheet, key, lang)
        col = get_index_from_language_list(lang)
        if rval != "":
            if is_first:
                is_first = False
                export_sheet.cell(row=export_row_index+1, column=1, value=key)
            export_sheet.cell(row=export_row_index+1, column=col+1, value=rval)
    export_row_index += 1


def read_raw_language_list():
    table = xlrd.open_workbook(raw_path)
    sheet = table.sheet_by_name("languageList")
    nrow = sheet.nrows
    col = get_cell_index_by_name(sheet, 1, "language")
    for row in range(4, nrow):
        languageList.append(sheet.cell_value(row, col))


def create_export_xlsx():
    export_book = openpyxl.Workbook()
    export_sheet = export_book.create_sheet()
    export_sheet.title = "language"

    size = len(languageList) + 1
    firstRows = ["" for _ in range(size)]
    secondRows = [a for a in languageList]
    secondRows.insert(0, "")
    thirdRows = [i == 0 and "uid" or "string" for i in range(size)]
    fourRows = ["" for _ in range(size)]

    for col in range(1, size+1):
        export_sheet.cell(row=1, column=col, value=firstRows[col - 1])
        export_sheet.cell(row=2, column=col, value=secondRows[col - 1])
        export_sheet.cell(row=3, column=col, value=thirdRows[col - 1])
        export_sheet.cell(row=4, column=col, value=fourRows[col - 1])
    return export_book, export_sheet


def main():
    read_raw_language_list()
    lang_table = xlrd.open_workbook(lang_path)
    lang_sheet = lang_table.sheet_by_name("language")
    print("languageList", languageList)
    # time.sleep(2)

    export_book, export_sheet = create_export_xlsx()

    for path, dir, filelist in os.walk(lang_dir):
        for filename in filelist:
            name = re.findall(r"^[^~$]*.xlsx$", filename)
            if name and name[0]:
                xlspath = os.path.join(path, filename)
                raw_book = None
                print("start: ")
                print(xlspath)
                raw_table = xlrd.open_workbook(xlspath)
                sheetname = "language"
                raw_sheet = raw_table.sheet_by_name(sheetname)

                global export_row_index

                for row in range(4, raw_sheet.nrows):
                    key = raw_sheet.cell_value(row, 0)
                    if key != "":
                        if is_key_diff(raw_sheet, lang_sheet, key):
                            export_diff_key(export_sheet, raw_sheet, key)
                            # raw_book = raw_key_fill_color(raw_sheet, key, xlspath)
                        else:
                            sync_equal_key(raw_sheet, lang_sheet, key, xlspath)
                else:
                    write_equal_list(xlspath, sheetname)
                    if raw_book:
                        raw_book.save(xlspath)
                print("end: ")
                print(filename)

    export_book.save(export_path)


if __name__ == "__main__":
    main()
