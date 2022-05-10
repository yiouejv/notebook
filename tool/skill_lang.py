#!/bin/python
# coding=utf-8
#Author: yiouejv
#Email: yiouejv@126.com
#Time: 2022-05-09 18:02:04
#Name: skill_lang.py

'''
    从filepath 读取 sheetname 页签,
    将 "技能描述" 这一列导出到 langSkillPath
    并添加新列 "技能desc", 将该列每行的内容填充为对应的key
'''

import xlrd
import openpyxl

filepath = u"xls/J技能表.xlsx"
sheetname = "skill"
langSkillPath = u"xls/lang/J技能表-多语言.xlsx"
typeRow = 2  # 从1开始, openpyxl


def getCellIndexByName(sheet, rowIndex, colname):
    ncol = sheet.ncols
    for i in range(0, ncol):
        v = sheet.cell_value(rowIndex, i)
        if v == colname:
            return i
    return -1

def getWaitTransCol(sheet, colIndex):
    index = colIndex
    nrows = sheet.nrows
    print("==============>total {} row".format(nrows), nrows)
    row2value = {}
    for row in range(4, nrows):
        v = sheet.cell_value(row, index)
        row2value[row + 1] = v
        # print("row {}:".format(row + 1), v)
    return row2value

def getTextId(row):
    return "skill_config__skill_text_{}".format(row)

def writeLanguage(book, row2value):
    # openpyxl 读写单元格时，单元格的坐标位置起始值是（1,1），即下标最小值为1，否则报错！
    wsheet = book.create_sheet()
    wsheet.title = "language"

    firstRows = [u"文本编号", u"中文"]
    secondRows = ["textId", "zh"]
    thirdRows = ["string", "string"]
    fourRows = ["" for _ in range(len(firstRows))]

    for col in range(1, len(firstRows) + 1):
        wsheet.cell(row=1, column=col, value=firstRows[col - 1])
        wsheet.cell(row=2, column=col, value=secondRows[col - 1])
        wsheet.cell(row=3, column=col, value=thirdRows[col - 1])
        wsheet.cell(row=4, column=col, value=fourRows[col - 1])

    rowIndex = 5
    for row, value in row2value.items():
        if value != "":
            textId = getTextId(row)
            wsheet.cell(rowIndex, column=1, value=textId)
            wsheet.cell(rowIndex, column=2, value=value)
            rowIndex += 1

    writeLanguageList(book)

    book.save(langSkillPath)

def writeLanguageList(book):
    sheet = book.create_sheet()
    sheet.title = "languageList"

    firstRows = [u"国家", u"语言", u"语言显示"]
    secondRows = [u"nation", "language", "show"]
    thirdRows = [u"string", "string", "string"]
    fourRows = ["" for _ in range(len(firstRows))]
    fiveRows = [u"中国", "zh", u"中文"]
    sixRows = [u"英国", "en", "English"]
    sevenRows = [u"德国", "de", "English"]
    eightRows = [u"法国", "fr", "English"]
    nineRows = [u"日本", "dex", "English"]
    tenRows = [u"韩国", "ko", "English"]

    for col in range(1, len(firstRows) + 1):
        sheet.cell(row=1, column=col, value=firstRows[col - 1])
        sheet.cell(row=2, column=col, value=secondRows[col - 1])
        sheet.cell(row=3, column=col, value=thirdRows[col - 1])
        sheet.cell(row=4, column=col, value=fourRows[col - 1])
        sheet.cell(row=5, column=col, value=fiveRows[col - 1])
        sheet.cell(row=6, column=col, value=sixRows[col - 1])
        sheet.cell(row=7, column=col, value=sevenRows[col - 1])
        sheet.cell(row=8, column=col, value=eightRows[col - 1])
        sheet.cell(row=9, column=col, value=nineRows[col - 1])
        sheet.cell(row=10, column=col, value=tenRows[col - 1])

def writeMain(table):
    sheet = table.active
    sheet.title = "main"
    firstRows = [u"子表", "", ""]
    secondRows = ["", u"//描述", u"//备注"]
    thirdRows = ["uid", "", ""]
    fourRows = ["" for _ in range(len(firstRows))]
    fiveRows = ["language", u"多语言配置表", ""]
    sixRows = ["languageList", u"多语言列表", ""]
    for col in range(1, len(firstRows) + 1):
        sheet.cell(row=1, column=col, value=firstRows[col - 1])
        sheet.cell(row=2, column=col, value=secondRows[col - 1])
        sheet.cell(row=3, column=col, value=thirdRows[col - 1])
        sheet.cell(row=4, column=col, value=fourRows[col - 1])
        sheet.cell(row=5, column=col, value=fiveRows[col - 1])
        sheet.cell(row=6, column=col, value=sixRows[col - 1])

def writeSkillDescCol(row2value):
    book = openpyxl.load_workbook(filepath)
    sheet = None
    for shet in book.worksheets:
        if shet.title == sheetname:
            sheet = shet
            break
    if not sheet:
        raise Exception("not find skill sheet")

    descColIndex = None
    ncols = sheet.max_column
    for colIdx in range(1, ncols + 1):
        col = sheet.cell(row=typeRow, column=colIdx)
        if col.value == "desc":
            descColIndex = colIdx
            break

    if descColIndex is None:
        # 不存在，插入一列
        descColIndex = ncols+ 1
        sheet.insert_cols(descColIndex)

    sheet.cell(row=1, column=descColIndex, value=u"技能desc")
    sheet.cell(row=2, column=descColIndex, value="desc")
    sheet.cell(row=3, column=descColIndex, value="string")
    sheet.cell(row=4, column=descColIndex, value="c")

    # 存在，修改desc列
    for row, value in row2value.items():
        textid = value != "" and getTextId(row) or ""
        sheet.cell(row=row, column=descColIndex, value=textid)
        print("change skill desc key, row={}, key={}".format(row, textid))

    book.save(filepath)


def main():
    table = xlrd.open_workbook(filepath)
    sheet = table.sheet_by_name(sheetname)
    index = getCellIndexByName(sheet, 0, u"技能描述")
    if index == -1:
        raise Exception(u"技能描述列没有找到")
    row2value = getWaitTransCol(sheet, index)

    # 导出多语言表
    book = openpyxl.Workbook()
    writeMain(book)
    writeLanguage(book, row2value)

    # 修改技能表的描述列
    writeSkillDescCol(row2value)


if __name__ == "__main__":
    main()
